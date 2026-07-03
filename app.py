#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, os, webbrowser, threading, time, hashlib, socket

_BASE   = os.path.dirname(os.path.abspath(__file__))
_WHEELS = os.path.join(_BASE, 'wheels')
if os.path.isdir(_WHEELS) and _WHEELS not in sys.path:
    sys.path.insert(0, _WHEELS)

import sqlite3, json, io, datetime
from flask import Flask, request, jsonify, send_file, g, session
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(_BASE, 'data', 'teammanager.db')
STATIC  = os.path.join(_BASE, 'static')
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

def get_local_ipv4():
    """获取本机局域网 IPv4 地址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

LOCAL_IPV4 = get_local_ipv4()  # 启动时获取一次，如 10.21.230.243

app = Flask(__name__, static_folder=STATIC, static_url_path='/static')
app.secret_key = 'team-mgr-2025-v3'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # disable static file cache

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    db.executescript("""
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;
CREATE TABLE IF NOT EXISTS members (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL, username TEXT UNIQUE NOT NULL,
    password TEXT NOT NULL DEFAULT '',
    email TEXT, phone TEXT, role TEXT DEFAULT 'DEVELOPER',
    group_name TEXT, is_admin INTEGER DEFAULT 0,
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS check_ins (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    member_id INTEGER NOT NULL REFERENCES members(id),
    check_date TEXT NOT NULL, check_in_time TEXT,
    status TEXT DEFAULT 'PRESENT', remark TEXT,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(member_id, check_date)
);
CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    title TEXT NOT NULL, description TEXT,
    task_type TEXT DEFAULT 'REQUIREMENT',
    status TEXT DEFAULT 'PENDING', priority TEXT DEFAULT 'MEDIUM',
    severity TEXT, issue_type TEXT,
    assignee_id INTEGER, assignee_name TEXT,
    reporter_name TEXT, group_name TEXT,
    plan_start_date TEXT, plan_end_date TEXT, actual_end_date TEXT,
    delivery_month TEXT, progress INTEGER DEFAULT 0,
    has_risk INTEGER DEFAULT 0, risk_description TEXT,
    version TEXT, module TEXT, location TEXT,
    estimated_days INTEGER DEFAULT 0,
    parent_task_id INTEGER,
    created_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS task_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    member_id INTEGER NOT NULL, member_name TEXT,
    log_date TEXT NOT NULL, content TEXT NOT NULL,
    progress_snapshot INTEGER, status_snapshot TEXT,
    hours REAL DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS system_config (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL,
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ci ON check_ins(member_id, check_date);
CREATE INDEX IF NOT EXISTS idx_tg ON tasks(group_name);
CREATE INDEX IF NOT EXISTS idx_ta ON tasks(assignee_id);
CREATE INDEX IF NOT EXISTS idx_tt ON tasks(task_type);
CREATE INDEX IF NOT EXISTS idx_lt ON task_logs(task_id);
CREATE INDEX IF NOT EXISTS idx_lm ON task_logs(member_id);
""")
    # Migration: add estimated_days if not exists
    try:
        db.execute("ALTER TABLE tasks ADD COLUMN estimated_days INTEGER DEFAULT 0")
        db.commit()
    except: pass
    try:
        db.execute("ALTER TABLE tasks ADD COLUMN parent_task_id INTEGER")
        db.commit()
    except: pass
    # Migration: add ip_address to check_ins
    try:
        db.execute("ALTER TABLE check_ins ADD COLUMN ip_address TEXT")
        db.commit()
    except: pass
    try:
        db.execute("ALTER TABLE task_logs ADD COLUMN hours REAL DEFAULT 0")
        db.commit()
    except: pass
    try:
        admin_pw = hash_pw('admin123')
        default_pw = hash_pw('123456')
        db.execute("INSERT OR IGNORE INTO members(name,username,password,role,group_name,is_admin) VALUES(?,?,?,?,?,?)",
                   ('管理员','admin',admin_pw,'ADMIN','研发一组',1))
        # Migration: if admin group_name is empty, set it to the first active non-admin member's group
        admin_row = db.execute("SELECT group_name FROM members WHERE username='admin'").fetchone()
        if not admin_row or not admin_row[0]:
            first_group = db.execute("SELECT group_name FROM members WHERE is_active=1 AND is_admin=0 AND group_name!='' LIMIT 1").fetchone()
            if first_group:
                db.execute("UPDATE members SET group_name=? WHERE username='admin'", (first_group[0],))
        for row in [('张三','zhangsan','研发一组','LEADER'),
                    ('李四','lisi','研发一组','DEVELOPER'),
                    ('王五','wangwu','研发一组','DEVELOPER'),
                    ('赵六','zhaoliu','研发一组','TESTER')]:
            db.execute("INSERT OR IGNORE INTO members(name,username,password,role,group_name,is_admin) VALUES(?,?,?,?,?,0)",
                       (row[0],row[1],default_pw,row[3],row[2]))
        db.commit()
        # default late threshold: 09:00
        db.execute("INSERT OR IGNORE INTO system_config(key,value) VALUES('late_threshold','09:00')")
        db.commit()
    except: pass
    db.close()

def r2d(r): return dict(r) if r else None
def rs(rows): return [dict(r) for r in rows]
def today(): return datetime.date.today().isoformat()
def now_str(): return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def workdays(year, month):
    import calendar
    return sum(1 for d in range(1, calendar.monthrange(year,month)[1]+1)
               if datetime.date(year,month,d).weekday()<5)

def auto_risk(d):
    done = ('DELIVERED','COMPLETED','RESOLVED','CLOSED','REJECTED','CANCELLED')
    if d.get('status') in done: return d
    t=today(); pe=d.get('plan_end_date') or ''; prog=int(d.get('progress',0) or 0)
    if pe and pe < t:
        d['has_risk']=1
        if not d.get('risk_description'): d['risk_description']='已超过计划截止日期'
    elif pe:
        warn=(datetime.date.today()+datetime.timedelta(days=7)).isoformat()
        if pe<=warn and prog<80:
            d['has_risk']=1
            if not d.get('risk_description'): d['risk_description']='临近截止，进度不足80%'
    return d

STATUS_ZH={'PENDING':'待处理','IN_PROGRESS':'进行中','TESTING':'测试中',
           'DELIVERED':'已交付','CANCELLED':'已取消','OPEN':'待处理',
           'RESOLVED':'已解决','CLOSED':'已关闭','REJECTED':'已拒绝',
           'ONGOING':'进行中','COMPLETED':'已完成'}
TYPE_ZH={'REQUIREMENT':'需求','ISSUE':'问题单','ONSITE':'现场支撑','OTHER':'其他事务'}

def current_user():
    uid=session.get('user_id')
    if not uid: return None
    row=get_db().execute("SELECT * FROM members WHERE id=? AND is_active=1",(uid,)).fetchone()
    return r2d(row)

def login_required(f):
    from functools import wraps
    @wraps(f)
    def dec(*a,**k):
        if not session.get('user_id'): return jsonify({'error':'未登录','code':401}),401
        return f(*a,**k)
    return dec

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def dec(*a,**k):
        u=current_user()
        if not u or not u.get('is_admin'): return jsonify({'error':'权限不足','code':403}),403
        return f(*a,**k)
    return dec

@app.after_request
def cors(r):
    r.headers['Access-Control-Allow-Origin']='*'
    r.headers['Access-Control-Allow-Methods']='GET,POST,PUT,DELETE,OPTIONS'
    r.headers['Access-Control-Allow-Headers']='Content-Type'
    return r

@app.route('/api/<path:p>', methods=['OPTIONS'])
def opt(p): return '',204

@app.route('/')
def spa():
    from flask import make_response
    import time
    resp = make_response(send_file(os.path.join(STATIC,'index.html')))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    resp.headers['ETag'] = str(int(time.time()))
    return resp

# ── Auth ─────────────────────────────────────────────────────
@app.route('/api/auth/login', methods=['POST'])
def login():
    d=request.json or {}
    row=get_db().execute("SELECT * FROM members WHERE username=? AND is_active=1",(d.get('username','').strip(),)).fetchone()
    if not row: return jsonify({'error':'用户名不存在'}),401
    m=r2d(row)
    raw_pw = d.get('password','')
    already_hashed = d.get('hashed', False)
    # 兼容：前端传来的是 sha256(明文)，直接比对
    # 也兼容旧客户端直接传明文（再做一次hash）
    if already_hashed:
        pw_match = (m['password'] == raw_pw)
    else:
        pw_match = (m['password'] == hash_pw(raw_pw))
    if not pw_match: return jsonify({'error':'密码错误'}),401
    session['user_id']=m['id']; m.pop('password',None)
    return jsonify({'user':m})

@app.route('/api/auth/logout', methods=['POST'])
def logout(): session.clear(); return jsonify({'ok':True})

@app.route('/api/auth/me')
def me():
    u=current_user()
    if not u: return jsonify({'error':'未登录','code':401}),401
    u.pop('password',None); return jsonify({'user':u})

@app.route('/api/auth/change_password', methods=['POST'])
@login_required
def change_password():
    d=request.json or {}
    old_pw=d.get('old_password',''); new_pw=d.get('new_password','')
    already_hashed=d.get('hashed',False)
    db=get_db(); uid=session['user_id']
    row=db.execute("SELECT password FROM members WHERE id=?",(uid,)).fetchone()
    if not row: return jsonify({'error':'用户不存在'}),400
    # 验证旧密码
    if already_hashed:
        ok = (row['password'] == old_pw)
    else:
        ok = (row['password'] == hash_pw(old_pw))
    if not ok: return jsonify({'error':'原密码错误'}),400
    # 新密码存储：如果前端已hash直接存，否则再hash
    if already_hashed:
        if len(new_pw) != 64: return jsonify({'error':'密码格式错误'}),400
        stored_pw = new_pw
    else:
        if len(new_pw)<6: return jsonify({'error':'密码至少6位'}),400
        stored_pw = hash_pw(new_pw)
    db.execute("UPDATE members SET password=? WHERE id=?",(stored_pw,uid)); db.commit()
    return jsonify({'ok':True})

# ── Members ──────────────────────────────────────────────────
@app.route('/api/members')
@login_required
def list_members():
    u=current_user(); db=get_db()
    if u['is_admin']:
        rows=db.execute("SELECT id,name,username,email,phone,role,group_name,is_admin,is_active,created_at FROM members ORDER BY id").fetchall()
    else:
        rows=db.execute("SELECT id,name,username,email,phone,role,group_name,is_admin,is_active,created_at FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(u['group_name'],)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/members/active')
@login_required
def active_members():
    u=current_user(); db=get_db()
    if u['is_admin']:
        rows=db.execute("SELECT id,name,username,role,group_name,is_admin FROM members WHERE is_active=1 ORDER BY id").fetchall()
    else:
        rows=db.execute("SELECT id,name,username,role,group_name,is_admin FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(u['group_name'],)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/members', methods=['POST'])
@admin_required
def add_member():
    d=request.json; db=get_db()
    raw=d.get('password','123456'); pw_store=hash_pw(raw) if d.get('pw_plain') else (raw if len(raw)==64 else hash_pw(raw))
    c=db.execute("INSERT INTO members(name,username,password,email,phone,role,group_name,is_admin,is_active) VALUES(?,?,?,?,?,?,?,?,?)",
        (d['name'],d['username'],pw_store,d.get('email'),d.get('phone'),
         d.get('role','DEVELOPER'),d.get('group_name'),1 if d.get('is_admin') else 0,1 if d.get('is_active',True) else 0))
    db.commit()
    row=db.execute("SELECT id,name,username,email,phone,role,group_name,is_admin,is_active FROM members WHERE id=?",(c.lastrowid,)).fetchone()
    return jsonify(r2d(row)),201

@app.route('/api/members/<int:mid>', methods=['PUT'])
@login_required
def upd_member(mid):
    u=current_user(); d=request.json; db=get_db()
    if not u['is_admin']:
        if mid!=u['id']: return jsonify({'error':'无权限'}),403
        db.execute("UPDATE members SET email=?,phone=? WHERE id=?",(d.get('email'),d.get('phone'),mid))
    else:
        db.execute("UPDATE members SET name=?,username=?,email=?,phone=?,role=?,group_name=?,is_admin=?,is_active=? WHERE id=?",
            (d['name'],d['username'],d.get('email'),d.get('phone'),d.get('role','DEVELOPER'),
             d.get('group_name'),1 if d.get('is_admin') else 0,1 if d.get('is_active',True) else 0,mid))
        if d.get('password'):
            raw2=d['password']; pw2=hash_pw(raw2) if d.get('pw_plain') else (raw2 if len(raw2)==64 else hash_pw(raw2))
            db.execute("UPDATE members SET password=? WHERE id=?",(pw2,mid))
    db.commit()
    row=db.execute("SELECT id,name,username,email,phone,role,group_name,is_admin,is_active FROM members WHERE id=?",(mid,)).fetchone()
    return jsonify(r2d(row)) if row else ('',404)

@app.route('/api/members/<int:mid>', methods=['DELETE'])
@admin_required
def del_member(mid):
    u=current_user(); db=get_db()
    target=r2d(db.execute("SELECT * FROM members WHERE id=?",(mid,)).fetchone())
    if not target: return '',404
    # 只有 username='admin' 的超级管理员不可操作；普通管理员可被停用
    if target.get('username')=='admin': return jsonify({'error':'超级管理员账号不可停用'}),403
    # 非超级管理员不能操作同级管理员（可选：这里允许任意管理员互相停用）
    db.execute("UPDATE members SET is_active=0 WHERE id=?",(mid,)); db.commit(); return '',204

@app.route('/api/members/<int:mid>/delete', methods=['DELETE'])
@admin_required
def hard_del_member(mid):
    """Hard delete a member - removes all their data."""
    db=get_db()
    target=r2d(db.execute("SELECT * FROM members WHERE id=?",(mid,)).fetchone())
    if not target: return '',404
    if target.get('username')=='admin': return jsonify({'error':'超级管理员账号不可删除'}),403
    # 删除关联数据
    db.execute("DELETE FROM check_ins WHERE member_id=?",(mid,))
    db.execute("DELETE FROM task_logs WHERE member_id=?",(mid,))
    # 任务：把 assignee 清空，不删任务本身
    db.execute("UPDATE tasks SET assignee_id=NULL,assignee_name='（已删除）' WHERE assignee_id=?",(mid,))
    db.execute("DELETE FROM members WHERE id=?",(mid,))
    db.commit()
    return jsonify({'ok':True})

# ── Check-in ─────────────────────────────────────────────────
def get_config(key, default=''):
    row = get_db().execute("SELECT value FROM system_config WHERE key=?",(key,)).fetchone()
    return row['value'] if row else default

def set_config(key, value):
    db = get_db()
    db.execute("INSERT INTO system_config(key,value,updated_at) VALUES(?,?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at",
               (key, value, now_str()))
    db.commit()

def detect_status(check_time_str):
    """Auto-detect PRESENT vs LATE based on late_threshold config. check_time_str is HH:MM:SS or HH:MM"""
    threshold = get_config('late_threshold', '09:00')
    try:
        ct = check_time_str[:5]   # HH:MM
        return 'LATE' if ct > threshold else 'PRESENT'
    except:
        return 'PRESENT'

# ── System config ─────────────────────────────────────────────
@app.route('/api/config', methods=['GET'])
@login_required
def get_all_config():
    rows = get_db().execute("SELECT key,value FROM system_config").fetchall()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/config', methods=['POST'])
@admin_required
def update_config():
    d = request.json or {}
    for key, value in d.items():
        set_config(key, str(value))
    return jsonify({'ok': True})

# ── Check-in ─────────────────────────────────────────────────
@app.route('/api/myip')
def my_ip():
    """返回服务器看到的客户端 IP（局域网直连时即为真实 IPv4）"""
    ip = (request.headers.get('X-Forwarded-For','').split(',')[0].strip()
          or request.headers.get('X-Real-IP','')
          or request.remote_addr or '')
    # 剥离 IPv6 映射前缀
    if ip.startswith('::ffff:'):
        ip = ip[7:]
    return jsonify({'ip': ip})


@app.route('/api/checkin', methods=['POST'])
@login_required
def checkin():
    u = current_user(); d = request.json or {}
    mid = int(d.get('memberId', u['id']))
    is_admin = u['is_admin']
    # Members can only check in for themselves
    if not is_admin and mid != u['id']: mid = u['id']

    dt     = d.get('date', today())
    remark = d.get('remark', '')
    db     = get_db()

    # Determine check_in_time and status
    existing = r2d(db.execute("SELECT * FROM check_ins WHERE member_id=? AND check_date=?",(mid,dt)).fetchone())

    if is_admin:
        # Admin can set explicit time and status
        explicit_time   = d.get('check_in_time')   # HH:MM or None
        explicit_status = d.get('status')           # explicit override
        if explicit_time:
            ci_time = "{} {}".format(dt, explicit_time)
            status  = explicit_status or detect_status(explicit_time)
        elif explicit_status:
            # Admin sets status directly (e.g. ABSENT/LEAVE) without time
            ci_time = "{} {}".format(dt, now_str()[11:16]) if dt == today() else (existing or {}).get('check_in_time') or ''
            status  = explicit_status
        else:
            # 管理员自己签到（不传 status 也不传时间）：按当前时间自动判断是否迟到
            ci_time = now_str() if dt == today() else (existing or {}).get('check_in_time') or ''
            status  = explicit_status or detect_status(now_str()[11:16])
    else:
        # Member: time is always server-side NOW (only for today)
        if dt != today():
            # Member cannot check in for other dates
            return jsonify({'error': '只能为今天签到，历史日期请联系管理员修改'}), 403
        if existing:
            # Already checked in today — re-check keeps original time, just updates status/remark if needed
            ci_time = existing['check_in_time'] or now_str()
            # Use existing time to re-evaluate late status if no explicit status
            status  = d.get('status') or detect_status(ci_time[11:16] if ci_time else '00:00')
        else:
            ci_time = now_str()
            # Auto-detect PRESENT vs LATE; member cannot choose these two — only LEAVE/REMOTE/ABSENT allowed as override
            requested = d.get('status', '')
            if requested in ('LEAVE', 'REMOTE', 'ABSENT'):
                status = requested
            else:
                status = detect_status(ci_time[11:16])

    # IP：优先使用前端传来的（前端通过 /api/myip 获取），其次服务器检测
    client_ip = d.get('clientIp','').strip()
    if not client_ip:
        raw_ip = (request.headers.get('X-Forwarded-For','').split(',')[0].strip()
                  or request.headers.get('X-Real-IP','')
                  or request.remote_addr or '')
        if raw_ip.startswith('::ffff:'):
            raw_ip = raw_ip[7:]
        if raw_ip in ('::1', '127.0.0.1'):
            client_ip = LOCAL_IPV4
        else:
            client_ip = raw_ip

    db.execute("""INSERT INTO check_ins(member_id,check_date,check_in_time,status,remark,ip_address) VALUES(?,?,?,?,?,?)
        ON CONFLICT(member_id,check_date) DO UPDATE SET status=excluded.status,
        remark=excluded.remark,check_in_time=excluded.check_in_time,ip_address=excluded.ip_address""",
               (mid, dt, ci_time, status, remark, client_ip))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM check_ins WHERE member_id=? AND check_date=?",(mid,dt)).fetchone()))

@app.route('/api/checkin/bulk', methods=['POST'])
@admin_required
def bulk_checkin():
    """Admin bulk check-in: submit a list of {memberId, status, check_in_time, remark}"""
    d    = request.json or {}
    dt   = d.get('date', today())
    rows = d.get('records', [])   # [{memberId, status, check_in_time?, remark?}]
    db   = get_db()
    results = []
    for rec in rows:
        mid    = int(rec.get('memberId'))
        status = rec.get('status', 'PRESENT')
        ci_t   = rec.get('check_in_time')   # HH:MM
        remark = rec.get('remark', '')
        if ci_t:
            ci_time = "{} {}".format(dt, ci_t)
            # re-evaluate late only if status not explicitly set to LEAVE/ABSENT/REMOTE
            if status not in ('LEAVE','ABSENT','REMOTE'):
                status = detect_status(ci_t)
        else:
            ci_time = "{} 09:00:00".format(dt)
        db.execute("""INSERT INTO check_ins(member_id,check_date,check_in_time,status,remark) VALUES(?,?,?,?,?)
            ON CONFLICT(member_id,check_date) DO UPDATE SET status=excluded.status,
            remark=excluded.remark,check_in_time=excluded.check_in_time""",
                   (mid, dt, ci_time, status, remark))
        results.append({'memberId': mid, 'status': status})
    db.commit()
    return jsonify({'ok': True, 'count': len(results), 'records': results})

@app.route('/api/checkin/today')
@login_required
def checkin_today():
    u = current_user()
    row = get_db().execute("SELECT * FROM check_ins WHERE member_id=? AND check_date=?",(u['id'],today())).fetchone()
    return jsonify(r2d(row))

@app.route('/api/checkin/day/<dt>')
@admin_required
def checkin_day(dt):
    """Get all check-ins for a given date (admin only)."""
    gn  = request.args.get('group_name')
    db  = get_db()
    if gn:
        rows = rs(db.execute("""SELECT c.*,m.name as member_name,m.group_name FROM check_ins c
            JOIN members m ON c.member_id=m.id WHERE c.check_date=? AND m.group_name=? ORDER BY m.id""",(dt,gn)).fetchall())
    else:
        rows = rs(db.execute("""SELECT c.*,m.name as member_name,m.group_name FROM check_ins c
            JOIN members m ON c.member_id=m.id WHERE c.check_date=? ORDER BY m.id""",(dt,)).fetchall())
    return jsonify(rows)

@app.route('/api/checkin/summary/group/<gn>')
@login_required
def ci_summary(gn):
    import calendar
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    yr=int(request.args.get('year',datetime.date.today().year))
    mo=int(request.args.get('month',datetime.date.today().month))
    s="{}-{:02d}-01".format(yr,mo); e="{}-{:02d}-{:02d}".format(yr,mo,calendar.monthrange(yr,mo)[1])
    db=get_db()
    members=rs(db.execute("SELECT * FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(gn,)).fetchall())
    sums=[]
    for m in members:
        rows=rs(db.execute("SELECT * FROM check_ins WHERE member_id=? AND check_date>=? AND check_date<=? ORDER BY check_date",(m['id'],s,e)).fetchall())
        from collections import Counter; cnt=Counter(r['status'] for r in rows)
        sums.append({'memberId':m['id'],'memberName':m['name'],'presentDays':cnt.get('PRESENT',0),
            'absentDays':cnt.get('ABSENT',0),'lateDays':cnt.get('LATE',0),
            'leaveDays':cnt.get('LEAVE',0),'remoteDays':cnt.get('REMOTE',0),'checkIns':rows})
    return jsonify({'groupName':gn,'year':yr,'month':mo,'totalWorkDays':workdays(yr,mo),'memberCount':len(members),'members':sums})

@app.route('/api/checkin/summary/annual/<gn>')
@login_required
def ci_annual(gn):
    import calendar
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    yr=int(request.args.get('year',datetime.date.today().year))
    db=get_db()
    members=rs(db.execute("SELECT * FROM members WHERE group_name=? AND is_active=1",(gn,)).fetchall())
    monthly=[]
    for mo in range(1,13):
        s="{}-{:02d}-01".format(yr,mo); e="{}-{:02d}-{:02d}".format(yr,mo,calendar.monthrange(yr,mo)[1])
        sums=[]
        for mem in members:
            rows=rs(db.execute("SELECT * FROM check_ins WHERE member_id=? AND check_date>=? AND check_date<=?",(mem['id'],s,e)).fetchall())
            from collections import Counter; cnt=Counter(r['status'] for r in rows)
            sums.append({'memberId':mem['id'],'memberName':mem['name'],'presentDays':cnt.get('PRESENT',0),'remoteDays':cnt.get('REMOTE',0),'absentDays':cnt.get('ABSENT',0),'checkIns':rows})
        monthly.append({'month':mo,'year':yr,'totalWorkDays':workdays(yr,mo),'memberCount':len(members),'members':sums})
    return jsonify({'groupName':gn,'year':yr,'monthlyData':monthly})


@app.route('/api/checkin/member/<int:mid>')
@admin_required
def checkin_member_records(mid):
    """Get all checkin records for a member in a given year/month."""
    from datetime import datetime as dt_cls
    import calendar
    year=int(request.args.get('year',dt_cls.now().year))
    month=int(request.args.get('month',dt_cls.now().month))
    s=f'{year}-{month:02d}-01'
    last_day=calendar.monthrange(year,month)[1]
    e=f'{year}-{month:02d}-{last_day:02d}'
    rows=get_db().execute(
        "SELECT * FROM check_ins WHERE member_id=? AND check_date>=? AND check_date<=? ORDER BY check_date",
        (mid,s,e)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/checkin/today_all')
@admin_required
def checkin_today_all():
    """Get today's check-in for all members in a group (admin only)."""
    gn = request.args.get('group_name', current_user()['group_name'])
    db = get_db()
    members = rs(db.execute("SELECT id,name,role FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(gn,)).fetchall())
    today_str = today()
    result = []
    for m in members:
        ci = r2d(db.execute("SELECT * FROM check_ins WHERE member_id=? AND check_date=?",(m['id'],today_str)).fetchone())
        result.append({
            'memberId': m['id'],
            'memberName': m['name'],
            'role': m['role'],
            'status': ci['status'] if ci else None,
            'check_in_time': ci['check_in_time'] if ci else None,
            'remark': ci['remark'] if ci else None,
            'ip_address': ci['ip_address'] if ci else None,
        })
    return jsonify(result)

# ── Tasks ────────────────────────────────────────────────────
def task_where(u, extra='', extra_params=()):
    if u['is_admin']: w,p="1=1",[]
    else: w,p="group_name=?",[u['group_name']]
    if extra: w+=" AND "+extra; p+=list(extra_params)
    return w,p

def can_edit(u,task):
    if u['is_admin']: return True
    return task.get('assignee_id')==u['id'] or task.get('created_by')==u['id']

@app.route('/api/tasks')
@login_required
def all_tasks():
    u=current_user(); db=get_db()
    tt=request.args.get('type')
    w,p=task_where(u,"task_type=?",(tt,)) if tt else task_where(u)
    rows=db.execute("SELECT * FROM tasks WHERE {} ORDER BY updated_at DESC".format(w),p).fetchall()
    return jsonify(rs(rows))

@app.route('/api/tasks/mine')
@login_required
def my_tasks():
    u=current_user(); db=get_db()
    tt=request.args.get('type')
    if tt:
        rows=db.execute("SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?) AND task_type=? ORDER BY updated_at DESC",(u['id'],u['id'],tt)).fetchall()
    else:
        rows=db.execute("SELECT * FROM tasks WHERE assignee_id=? OR created_by=? ORDER BY updated_at DESC",(u['id'],u['id'])).fetchall()
    return jsonify(rs(rows))

@app.route('/api/tasks/risk')
@login_required
def risk_tasks():
    u=current_user(); db=get_db(); t=today()
    w,p=task_where(u,"status NOT IN ('DELIVERED','CANCELLED','RESOLVED','CLOSED','REJECTED','COMPLETED') AND (plan_end_date<? OR has_risk=1)",(t,))
    return jsonify(rs(db.execute("SELECT * FROM tasks WHERE {} ORDER BY plan_end_date".format(w),p).fetchall()))

@app.route('/api/tasks/gantt')
@login_required
def gantt_tasks():
    u=current_user(); db=get_db()
    mid=request.args.get('member_id',u['id'])
    yr=int(request.args.get('year',datetime.date.today().year))
    mo=request.args.get('month')
    if not u['is_admin'] and str(mid)!=str(u['id']): mid=u['id']
    if mo:
        import calendar as cal
        s="{}-{:02d}-01".format(yr,int(mo))
        e="{}-{:02d}-{:02d}".format(yr,int(mo),cal.monthrange(yr,int(mo))[1])
    else:
        s="{}-01-01".format(yr); e="{}-12-31".format(yr)
    rows=db.execute("""SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?)
        AND plan_start_date IS NOT NULL AND plan_end_date IS NOT NULL
        AND plan_start_date<=? AND plan_end_date>=?
        ORDER BY plan_start_date""",(mid,mid,e,s)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/tasks/stats/<gn>')
@login_required
def task_stats(gn):
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    s=request.args.get('startMonth',''); e=request.args.get('endMonth','')
    tt=request.args.get('type','REQUIREMENT')
    rows=rs(get_db().execute("SELECT delivery_month,status,COUNT(*) as cnt FROM tasks WHERE group_name=? AND task_type=? AND delivery_month>=? AND delivery_month<=? GROUP BY delivery_month,status ORDER BY delivery_month",(gn,tt,s,e)).fetchall())
    monthly={}
    for r in rows: monthly.setdefault(r['delivery_month'] or 'unknown',{})[r['status']]=r['cnt']
    return jsonify({'groupName':gn,'startMonth':s,'endMonth':e,'monthlyStats':monthly,'taskType':tt})

@app.route('/api/tasks', methods=['POST'])
@login_required
def add_task():
    u=current_user(); d=dict(request.json)
    if not d.get('group_name'): d['group_name']=u['group_name']
    d['created_by']=u['id']
    # Auto-assign to self if not specified
    if not d.get('assignee_id'): d['assignee_id']=u['id']; d['assignee_name']=u['name']
    if not d.get('reporter_name'): d['reporter_name']=u['name']
    if not d.get('delivery_month') and d.get('plan_end_date'): d['delivery_month']=d['plan_end_date'][:7]
    d=auto_risk(d); db=get_db()
    c=db.execute("""INSERT INTO tasks(title,description,task_type,status,priority,severity,issue_type,
        assignee_id,assignee_name,reporter_name,group_name,plan_start_date,plan_end_date,actual_end_date,
        delivery_month,progress,has_risk,risk_description,version,module,location,estimated_days,parent_task_id,created_by)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (d.get('title'),d.get('description'),d.get('task_type','REQUIREMENT'),
         d.get('status','PENDING'),d.get('priority','MEDIUM'),d.get('severity'),d.get('issue_type'),
         d.get('assignee_id'),d.get('assignee_name'),d.get('reporter_name'),d.get('group_name'),
         d.get('plan_start_date'),d.get('plan_end_date'),d.get('actual_end_date'),
         d.get('delivery_month'),d.get('progress',0),d.get('has_risk',0),d.get('risk_description'),
         d.get('version'),d.get('module'),d.get('location'),int(d.get('estimated_days') or 0),d.get('parent_task_id'),d.get('created_by')))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM tasks WHERE id=?",(c.lastrowid,)).fetchone())),201

@app.route('/api/tasks/<int:tid>', methods=['GET'])
@login_required
def get_task(tid):
    row=get_db().execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone()
    return jsonify(r2d(row)) if row else ('',404)

@app.route('/api/tasks/<int:tid>', methods=['PUT'])
@login_required
def upd_task(tid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone())
    if not existing: return ('',404)
    if not can_edit(u,existing): return jsonify({'error':'只能修改自己创建或负责的任务'}),403
    d=dict(request.json)
    # Helper: 如果 d 里没传某字段（key缺失或值为None），fallback 用 existing 的值，避免误清空
    def keep(key, default=None):
        v = d.get(key)
        if v is None or v == '':
            return existing.get(key, default)
        return v
    if d.get('status') in ('DELIVERED','COMPLETED','RESOLVED','CLOSED') and not d.get('actual_end_date'):
        d['actual_end_date']=today()
        if not d.get('progress'): d['progress']=100
    d=auto_risk(d); db.execute("""UPDATE tasks SET title=?,description=?,task_type=?,status=?,priority=?,
        severity=?,issue_type=?,assignee_id=?,assignee_name=?,reporter_name=?,plan_start_date=?,
        plan_end_date=?,actual_end_date=?,delivery_month=?,progress=?,has_risk=?,risk_description=?,
        version=?,module=?,location=?,estimated_days=?,parent_task_id=?,updated_at=? WHERE id=?""",
        (keep('title'),d.get('description'),keep('task_type'),
         keep('status'),keep('priority'),d.get('severity'),d.get('issue_type'),
         keep('assignee_id'),keep('assignee_name'),d.get('reporter_name'),
         d.get('plan_start_date'),d.get('plan_end_date'),d.get('actual_end_date'),
         d.get('delivery_month'),d.get('progress',existing.get('progress',0)),
         d.get('has_risk',0),d.get('risk_description'),
         d.get('version'),d.get('module'),d.get('location'),
         int(d.get('estimated_days') or existing.get('estimated_days') or 0),
         keep('parent_task_id'),now_str(),tid))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone()))

@app.route('/api/tasks/<int:tid>', methods=['DELETE'])
@login_required
def del_task(tid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone())
    if not existing: return ('',404)
    if not u['is_admin'] and existing.get('created_by')!=u['id']: return jsonify({'error':'只能删除自己创建的任务'}),403
    db.execute("DELETE FROM tasks WHERE id=?",(tid,)); db.commit(); return '',204

# ── Task Logs ─────────────────────────────────────────────────
@app.route('/api/tasks/<int:tid>/logs')
@login_required
def get_logs(tid):
    rows=get_db().execute("SELECT * FROM task_logs WHERE task_id=? ORDER BY log_date DESC, id DESC",(tid,)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/tasks/<int:tid>/logs', methods=['POST'])
@login_required
def add_log(tid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone())
    if not existing: return ('',404)
    if not can_edit(u,existing): return jsonify({'error':'无权限'}),403
    d=request.json or {}
    content=d.get('content','').strip()
    if not content: return jsonify({'error':'进展内容不能为空'}),400
    log_date=d.get('log_date',today())
    new_prog=d.get('progress'); new_stat=d.get('status')
    try: hours=max(0.0,float(d.get('hours') or 0))
    except (TypeError,ValueError): hours=0.0
    if new_prog is not None or new_stat is not None:
        upd=[]; prm=[]
        if new_prog is not None: upd.append("progress=?"); prm.append(int(new_prog))
        if new_stat is not None: upd.append("status=?"); prm.append(new_stat)
        upd.append("updated_at=?"); prm.append(now_str()); prm.append(tid)
        db.execute("UPDATE tasks SET {} WHERE id=?".format(','.join(upd)),prm)
    c=db.execute("INSERT INTO task_logs(task_id,member_id,member_name,log_date,content,progress_snapshot,status_snapshot,hours) VALUES(?,?,?,?,?,?,?,?)",
        (tid,u['id'],u['name'],log_date,content,
         new_prog if new_prog is not None else existing.get('progress'),
         new_stat or existing.get('status'),hours))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM task_logs WHERE id=?",(c.lastrowid,)).fetchone())),201

@app.route('/api/tasks/logs/mine')
@login_required
def my_logs():
    u=current_user()
    days=int(request.args.get('days',7))
    since=(datetime.date.today()-datetime.timedelta(days=days)).isoformat()
    rows=get_db().execute("""SELECT l.*,t.title as task_title,t.task_type FROM task_logs l
        JOIN tasks t ON l.task_id=t.id WHERE l.member_id=? AND l.log_date>=?
        ORDER BY l.log_date DESC,l.id DESC LIMIT 50""",(u['id'],since)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/stats/delivery')
@login_required
def delivery_stats():
    """
    Delivery performance per member.
    Returns on-time rate, delivered count, late count, sorted by on_time_rate desc.
    member_ids: comma-separated, optional filter
    """
    u = current_user()
    gn = request.args.get('group_name', u.get('group_name',''))
    start = request.args.get('startMonth', '')
    end   = request.args.get('endMonth', '')
    member_ids_param = request.args.get('member_ids', '')  # "1,2,3" or ""

    db = get_db()
    # Base query: REQUIREMENT tasks in range
    if u['is_admin']:
        base = "task_type='REQUIREMENT' AND group_name=?"
        params = [gn]
    else:
        base = "task_type='REQUIREMENT' AND group_name=?"
        params = [gn]

    if start: base += " AND delivery_month>=?"; params.append(start)
    if end:   base += " AND delivery_month<=?"; params.append(end)

    tasks = rs(db.execute(f"SELECT * FROM tasks WHERE {base}", params).fetchall())

    # Filter by member_ids if provided
    selected_mids = set()
    if member_ids_param:
        selected_mids = set(int(x) for x in member_ids_param.split(',') if x.strip().isdigit())
        tasks = [t for t in tasks if t.get('assignee_id') in selected_mids]

    # Group by assignee
    from collections import defaultdict
    member_tasks = defaultdict(list)
    for t in tasks:
        mid = t.get('assignee_id')
        if mid:
            member_tasks[mid].append(t)

    results = []
    for mid, mtasks in member_tasks.items():
        name = next((t['assignee_name'] for t in mtasks if t['assignee_name']), str(mid))
        delivered = [t for t in mtasks if t['status'] == 'DELIVERED']
        total     = len(mtasks)
        d_count   = len(delivered)
        # On time = delivered AND actual_end_date <= plan_end_date
        on_time   = sum(1 for t in delivered
                        if t.get('actual_end_date') and t.get('plan_end_date')
                        and t['actual_end_date'] <= t['plan_end_date'])
        late_del  = d_count - on_time
        on_time_rate = round(on_time / d_count * 100, 1) if d_count > 0 else 0.0
        results.append({
            'memberId':     mid,
            'memberName':   name,
            'total':        total,
            'delivered':    d_count,
            'onTime':       on_time,
            'lateDelivery': late_del,
            'inProgress':   sum(1 for t in mtasks if t['status'] in ('IN_PROGRESS','TESTING','PENDING')),
            'onTimeRate':   on_time_rate,
        })

    # Sort by on_time_rate desc, then delivered desc
    results.sort(key=lambda x: (-x['onTimeRate'], -x['delivered']))
    return jsonify({'members': results, 'top3': results[:3]})

@app.route('/api/stats/timelog')
@login_required
def timelog_stats():
    """
    Per-member time allocation: hours spent per task, per task type, per day.
    Non-admins are locked to their own data; admins may pass any member_id.
    """
    u = current_user()
    member_id = request.args.get('member_id', type=int)
    if not u['is_admin']:
        member_id = u['id']
    if not member_id:
        return jsonify({'error':'member_id 不能为空'}), 400

    db = get_db()
    member = r2d(db.execute("SELECT id,name FROM members WHERE id=?", (member_id,)).fetchone())
    if not member: return jsonify({'error':'成员不存在'}), 404

    start = request.args.get('startMonth', '')
    end   = request.args.get('endMonth', '') or start

    q = """SELECT l.task_id,l.log_date,l.hours,t.title as task_title,t.task_type
           FROM task_logs l JOIN tasks t ON l.task_id=t.id WHERE l.member_id=?"""
    params = [member_id]
    if start: q += " AND l.log_date>=?"; params.append(start+'-01')
    if end:
        import calendar
        ey, em = map(int, end.split('-'))
        q += " AND l.log_date<=?"; params.append("{}-{:02d}".format(end, calendar.monthrange(ey,em)[1]))
    rows = rs(db.execute(q, params).fetchall())

    by_task = {}
    by_type = {}
    by_date = {}
    total_hours = 0.0
    for r in rows:
        h = r.get('hours') or 0.0
        total_hours += h
        tid = r['task_id']
        if tid not in by_task:
            by_task[tid] = {'taskId': tid, 'title': r['task_title'], 'taskType': r['task_type'], 'hours': 0.0, 'logCount': 0}
        by_task[tid]['hours'] += h
        by_task[tid]['logCount'] += 1
        by_type[r['task_type']] = by_type.get(r['task_type'], 0.0) + h
        by_date[r['log_date']] = by_date.get(r['log_date'], 0.0) + h

    task_list = sorted(by_task.values(), key=lambda x: -x['hours'])
    for t in task_list: t['hours'] = round(t['hours'], 2)
    by_type = {k: round(v, 2) for k, v in by_type.items()}
    by_date = {k: round(v, 2) for k, v in by_date.items()}

    total_wd = 0
    if start:
        import calendar
        sy, sm = map(int, start.split('-'))
        ey, em = map(int, end.split('-'))
        y, m = sy, sm
        while (y, m) <= (ey, em):
            total_wd += workdays(y, m)
            m += 1
            if m > 12: m = 1; y += 1

    return jsonify({
        'memberId': member_id, 'memberName': member['name'],
        'startMonth': start, 'endMonth': end,
        'totalHours': round(total_hours, 2),
        'totalWorkDays': total_wd,
        'avgHoursPerWorkday': round(total_hours/total_wd, 2) if total_wd else 0,
        'byTask': task_list,
        'byType': by_type,
        'byDate': by_date,
    })

# ── Export ────────────────────────────────────────────────────
def mkhdr(ws,row,hdrs,wds):
    f=PatternFill("solid",fgColor="1F3864"); ft=Font(bold=True,color="FFFFFF",size=11); al=Alignment(horizontal="center",vertical="center")
    for i,h in enumerate(hdrs,1):
        c=ws.cell(row=row,column=i,value=h); c.fill=f; c.font=ft; c.alignment=al
        ws.column_dimensions[get_column_letter(i)].width=wds[i-1]

def rf(): return PatternFill("solid",fgColor="FFE0B2")

def title_cell(ws,text,cols):
    ws.merge_cells('A1:{}1'.format(get_column_letter(cols)))
    ws['A1'].value=text; ws['A1'].font=Font(bold=True,size=14); ws['A1'].alignment=Alignment(horizontal='center')
    ws.row_dimensions[1].height=28

@app.route('/api/export/checkin/<gn>')
@login_required
def exp_ci(gn):
    import calendar
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    yr=int(request.args.get('year',datetime.date.today().year))
    mo=int(request.args.get('month',datetime.date.today().month))
    s="{}-{:02d}-01".format(yr,mo); e="{}-{:02d}-{:02d}".format(yr,mo,calendar.monthrange(yr,mo)[1])
    db=get_db(); members=rs(db.execute("SELECT * FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(gn,)).fetchall())
    wd=workdays(yr,mo); wb=Workbook(); ws=wb.active; ws.title="{}年{}月签到".format(yr,mo)
    title_cell(ws,"{} {}年{}月 签到报表".format(gn,yr,mo),9)
    mkhdr(ws,3,['姓名','工作日','出勤','缺勤','迟到','请假','远程','出勤率','备注'],[14,8,8,8,8,8,8,10,20])
    for m in members:
        rows=rs(db.execute("SELECT status FROM check_ins WHERE member_id=? AND check_date>=? AND check_date<=?",(m['id'],s,e)).fetchall())
        from collections import Counter; cnt=Counter(r['status'] for r in rows)
        p=cnt.get('PRESENT',0); rv=cnt.get('REMOTE',0)
        ws.append([m['name'],wd,p,cnt.get('ABSENT',0),cnt.get('LATE',0),cnt.get('LEAVE',0),rv,
                   "{:.1f}%".format((p+rv)/wd*100) if wd else "0%",''])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name="{}_签到_{}.xlsx".format(gn,yr))

@app.route('/api/export/tasks/<gn>')
@login_required
def exp_tasks(gn):
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    tt=request.args.get('type','REQUIREMENT')
    s=request.args.get('startMonth',''); e=request.args.get('endMonth','')
    rows=rs(get_db().execute("SELECT * FROM tasks WHERE group_name=? AND task_type=? AND delivery_month>=? AND delivery_month<=? ORDER BY delivery_month",(gn,tt,s,e)).fetchall())
    tname=TYPE_ZH.get(tt,tt); wb=Workbook(); ws=wb.active; ws.title=tname+"报表"
    title_cell(ws,"{} {} ({} ~ {})".format(gn,tname,s,e),12)
    mkhdr(ws,3,['标题','类型','模块','版本','负责人','优先级','状态','进度','计划开始','计划结束','实际完成','风险'],
          [32,10,12,10,10,8,10,8,12,12,12,22])
    for r in rows:
        risk=('⚠ '+(r.get('risk_description') or '有风险')) if r.get('has_risk') else '正常'
        ws.append([r['title'],TYPE_ZH.get(r.get('task_type',''),''),r.get('module',''),r.get('version',''),
                   r.get('assignee_name',''),r.get('priority',''),STATUS_ZH.get(r.get('status',''),r.get('status','')),
                   "{}%".format(r.get('progress',0)),r.get('plan_start_date',''),r.get('plan_end_date',''),r.get('actual_end_date',''),risk])
        if r.get('has_risk'):
            for col in range(1,13): ws.cell(row=ws.max_row,column=col).fill=rf()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name="{}_{}_{}_{}.xlsx".format(gn,tname,s,e))



@app.route('/api/tasks/<int:tid>/subtasks')
@login_required
def get_subtasks(tid):
    """Get all subtasks of a parent task."""
    rows=get_db().execute("SELECT * FROM tasks WHERE parent_task_id=? ORDER BY created_at",(tid,)).fetchall()
    return jsonify(rs(rows))

@app.route('/api/tasks/by_member')
@login_required
def tasks_by_member():
    """Admin: get tasks grouped by selected members with daily logs."""
    u = current_user()
    if not u['is_admin']: return jsonify({'error':'无权限'}),403
    member_ids = request.args.get('member_ids','')
    date_filter = request.args.get('date')   # optional: YYYY-MM-DD
    gn = request.args.get('group_name', u['group_name'])
    db = get_db()
    mids = [int(x) for x in member_ids.split(',') if x.strip().isdigit()] if member_ids else []
    if not mids:
        rows = rs(db.execute("SELECT id,name FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(gn,)).fetchall())
        mids = [r['id'] for r in rows]
    result = []
    for mid in mids:
        member = r2d(db.execute("SELECT id,name,role,group_name FROM members WHERE id=?",(mid,)).fetchone())
        if not member: continue
        tasks = rs(db.execute("SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?) ORDER BY plan_start_date,created_at",(mid,mid)).fetchall())
        # attach logs
        for t in tasks:
            logs = rs(db.execute("SELECT * FROM task_logs WHERE task_id=? ORDER BY log_date DESC,id DESC",(t['id'],)).fetchall())
            t['logs'] = logs
            # check if today has a log
            t['has_today_log'] = any(l['log_date']==today() for l in logs)
            if date_filter:
                t['has_date_log'] = any(l['log_date']==date_filter for l in logs)
        result.append({'member': member, 'tasks': tasks})
    return jsonify(result)

@app.route('/api/tasks/gantt_multi')
@login_required
def gantt_multi():
    """Gantt for multiple members (admin view)."""
    u = current_user()
    if not u['is_admin']: return jsonify({'error':'无权限'}),403
    member_ids = request.args.get('member_ids','')
    yr  = int(request.args.get('year', datetime.date.today().year))
    mo  = request.args.get('month')
    gn  = request.args.get('group_name', u['group_name'])
    db  = get_db()
    mids = [int(x) for x in member_ids.split(',') if x.strip().isdigit()] if member_ids else []
    if not mids:
        rows = rs(db.execute("SELECT id FROM members WHERE group_name=? AND is_active=1",(gn,)).fetchall())
        mids = [r['id'] for r in rows]
    if mo:
        import calendar as cal
        s = "{}-{:02d}-01".format(yr,int(mo))
        e = "{}-{:02d}-{:02d}".format(yr,int(mo),cal.monthrange(yr,int(mo))[1])
    else:
        s = "{}-01-01".format(yr); e = "{}-12-31".format(yr)
    result = []
    for mid in mids:
        member = r2d(db.execute("SELECT id,name,role,group_name FROM members WHERE id=?",(mid,)).fetchone())
        if not member: continue
        tasks = rs(db.execute("""SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?)
            AND plan_start_date IS NOT NULL AND plan_end_date IS NOT NULL
            AND plan_start_date<=? AND plan_end_date>=?
            ORDER BY plan_start_date""",(mid,mid,e,s)).fetchall())
        for t in tasks:
            logs = rs(db.execute("SELECT log_date,content,progress_snapshot,member_name FROM task_logs WHERE task_id=? ORDER BY log_date",(t['id'],)).fetchall())
            t['logs'] = logs
            t['log_dates'] = [l['log_date'] for l in logs]
        result.append({'member': member, 'tasks': tasks})
    return jsonify({'members': result, 'startDate': s, 'endDate': e})

@app.route('/api/tasks/today_todo')
@login_required
def today_todo():
    """Tasks spanning today (active tasks where plan_start<=today<=plan_end)."""
    u = current_user(); db = get_db(); t = today()
    if u['is_admin']:
        where,params = task_where(u,
            "plan_start_date<=? AND plan_end_date>=? AND status NOT IN ('DELIVERED','COMPLETED','RESOLVED','CLOSED','CANCELLED','REJECTED')",
            (t,t))
    else:
        where = "(assignee_id=? OR created_by=?) AND plan_start_date<=? AND plan_end_date>=? AND status NOT IN ('DELIVERED','COMPLETED','RESOLVED','CLOSED','CANCELLED','REJECTED')"
        params = [u['id'],u['id'],t,t]
    rows = rs(db.execute("SELECT * FROM tasks WHERE {} ORDER BY priority DESC,plan_end_date".format(where),params).fetchall())
    for task in rows:
        task['has_today_log'] = db.execute("SELECT 1 FROM task_logs WHERE task_id=? AND log_date=?",(task['id'],t)).fetchone() is not None
    return jsonify(rows)

@app.route('/api/tasks/monthly_view')
@login_required
def monthly_view():
    """Monthly task view: tasks active in the month, with daily log counts."""
    u = current_user(); db = get_db()
    yr = int(request.args.get('year', datetime.date.today().year))
    mo = int(request.args.get('month', datetime.date.today().month))
    member_ids = request.args.get('member_ids','')
    import calendar as cal
    s = "{}-{:02d}-01".format(yr,mo)
    e = "{}-{:02d}-{:02d}".format(yr,mo,cal.monthrange(yr,mo)[1])
    days_in_month = cal.monthrange(yr,mo)[1]
    mids = [int(x) for x in member_ids.split(',') if x.strip().isdigit()] if member_ids else []
    if u['is_admin']:
        gn = request.args.get('group_name', u['group_name'])
        if not mids:
            rows = rs(db.execute("SELECT id FROM members WHERE group_name=? AND is_active=1",(gn,)).fetchall())
            mids = [r['id'] for r in rows]
        # Query by member IDs (more reliable than group_name on tasks table)
        if mids:
            placeholders = ','.join('?' * len(mids))
            where = "(assignee_id IN ({}) OR created_by IN ({})) AND status NOT IN ('CANCELLED','REJECTED')".format(placeholders, placeholders)
            params = mids + mids
        else:
            where = "group_name=? AND status NOT IN ('CANCELLED','REJECTED')"
            params = [gn]
    else:
        where = "(assignee_id=? OR created_by=?) AND status NOT IN ('CANCELLED','REJECTED')"
        params = [u['id'],u['id']]
    all_tasks = rs(db.execute("SELECT * FROM tasks WHERE {} ORDER BY assignee_name".format(where),params).fetchall())
    # No need to filter by mids again - already handled in WHERE clause
    # Filter: keep tasks that overlap with the month OR have logs in the month
    tasks = []
    for t in all_tasks:
        ps = t.get('plan_start_date') or ''
        pe = t.get('plan_end_date') or ''
        # 1. Overlaps with month range
        overlaps = (not ps or ps <= e) and (not pe or pe >= s)
        # 2. Or has logs in the month
        has_logs = db.execute("SELECT 1 FROM task_logs WHERE task_id=? AND log_date>=? AND log_date<=? LIMIT 1",
                              (t['id'], s, e)).fetchone() is not None
        if overlaps or has_logs:
            tasks.append(t)
    for task in tasks:
        logs = rs(db.execute("SELECT log_date,content,progress_snapshot,member_name FROM task_logs WHERE task_id=? AND log_date>=? AND log_date<=? ORDER BY log_date",(task['id'],s,e)).fetchall())
        by_date = {}
        for l in logs:
            by_date.setdefault(l['log_date'],[]).append(l)
        task['logs_by_date'] = by_date
    all_dates = ["{}-{:02d}-{:02d}".format(yr,mo,d) for d in range(1,days_in_month+1)]
    return jsonify({'tasks':tasks,'dates':all_dates,'year':yr,'month':mo})

if __name__=='__main__':
    port=int(os.environ.get('PORT',8080))
    init_db()
    def open_browser():
        time.sleep(1.5); webbrowser.open('http://127.0.0.1:{}'.format(port))
    threading.Thread(target=open_browser,daemon=True).start()
    print("="*50)
    print("  Team Manager v3")
    print("  http://127.0.0.1:{}".format(port))
    print("  admin/admin123 | zhangsan/123456")
    print("="*50)
    app.run(host='0.0.0.0',port=port,debug=False)
