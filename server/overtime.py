# -*- coding: utf-8 -*-
"""Overtime requests: CRUD, lock/unlock, and Excel export."""
import io, datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from .db import get_db, r2d, rs
from .utils import today, now_str, current_user, login_required, admin_required

overtime_bp = Blueprint('overtime', __name__)

# Column layout mirrors the corporate overtime-registration template
# (overtime_registration.xlsx): 10 columns, no extra fields.
_OT_NOTE = '绿色区域为制度区域，茶色区与为表单区域；*标记字段为必填项，请勿改动模板格式'
_OT_HEADERS = [
    '*工号(8位)\nEmployee ID(8bit)', '姓名\nEmployee Name',
    '*开始日期\nStart Date\n(YYYY/MM/DD)', '*开始时间\nStart Time\n(HH:MM)',
    '*结束日期\nEnd Date\n(YYYY/MM/DD)', '*结束时间\nEnd Time\n(HH:MM)',
    '休息开始时间\nBreak Start Time\n(HH:MM)', '休息结束时间\nBreak End Time\n(HH:MM)',
    '*加班类型\nOvertime Type', '加班理由',
]
_OT_WIDTHS = [18, 16, 18, 14, 18, 14, 18, 18, 38.06640625, 20]
# Maps our overtime_type value to the exact option text used by the corporate template's 加班类型 column
_OT_TYPE_MAP = {'转加班费': 'CN_Overtime/转加班费', '转调休': 'CN_Overtime_for_Replacement_Leave/转调休'}
_THIN = Side(style='thin')
_OT_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_OT_GREEN = 'FFC6EFCE'   # 制度区域 (system area, e.g. A1/row2 headers) - ARGB, FF alpha required or openpyxl writes it transparent
_OT_TAN = 'FFFCD5B5'     # 表单区域 (form area, e.g. B1:J1 note banner) - template's Accent6/Lighter-60% theme tint
_HDR_FONT_KW = dict(name='宋体', bold=True, size=11)
_HDR_INLINE_KW = dict(rFont='宋体', b=True, sz=11)


def _hdr_value(text):
    """Header cell value with the leading '*' (required-field marker) rendered in red, matching the template."""
    if not text.startswith('*'):
        return text
    return CellRichText(
        TextBlock(InlineFont(color='FFFF0000', **_HDR_INLINE_KW), '*'),
        TextBlock(InlineFont(**_HDR_INLINE_KW), text[1:]),
    )


def _ot_sheet(wb, title):
    ws = wb.create_sheet(title=title)
    green = PatternFill('solid', fgColor=_OT_GREEN)
    tan = PatternFill('solid', fgColor=_OT_TAN)
    hdr_font = Font(**_HDR_FONT_KW)
    note_font = Font(name='宋体', bold=False, size=11)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bottom_only = Border(bottom=_THIN)

    ws['A1'] = '说明'
    ws['A1'].fill = green; ws['A1'].font = hdr_font; ws['A1'].alignment = center_wrap; ws['A1'].border = _OT_BORDER

    ws.merge_cells('B1:I1')
    ws['B1'] = _OT_NOTE
    ws['B1'].fill = tan; ws['B1'].font = note_font; ws['B1'].alignment = Alignment(horizontal='left')
    for col in 'BCDEFGHI':
        ws[col+'1'].border = bottom_only
    ws['J1'].fill = tan; ws['J1'].font = note_font

    for i, h in enumerate(_OT_HEADERS, 1):
        c = ws.cell(row=2, column=i, value=_hdr_value(h))
        c.fill = green; c.font = hdr_font; c.alignment = center_wrap; c.border = _OT_BORDER
        ws.column_dimensions[get_column_letter(i)].width = _OT_WIDTHS[i-1]
    ws.row_dimensions[2].height = 52.05
    return ws


def _ot_row(ws, row_idx, values):
    font = Font(name='Arial', size=10)
    align = Alignment(horizontal='center', vertical='center')
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=i, value=v)
        c.font = font; c.alignment = align; c.border = _OT_BORDER


def _can_edit_overtime(u, row):
    return (not row['locked']) and (u['is_admin'] or row['member_id']==u['id'])


@overtime_bp.route('/api/overtime')
@login_required
def list_overtime():
    db=get_db()
    month=request.args.get('month') or today()[:7]
    rows=rs(db.execute("SELECT * FROM overtime_requests WHERE start_date LIKE ? ORDER BY start_date,id",(month+'%',)).fetchall())
    return jsonify(rows)


@overtime_bp.route('/api/overtime', methods=['POST'])
@login_required
def add_overtime():
    u=current_user(); d=request.json or {}; db=get_db()
    employee_no=d.get('employee_no') or u.get('employee_no')
    start_date=d.get('start_date'); start_time=d.get('start_time')
    end_date=d.get('end_date'); end_time=d.get('end_time')
    overtime_type=d.get('overtime_type')
    if not (employee_no and start_date and start_time and end_date and end_time and overtime_type):
        return jsonify({'error':'请填写完整的必填项'}),400
    if overtime_type not in ('转加班费','转调休'):
        return jsonify({'error':'加班类型不合法'}),400
    c=db.execute("""INSERT INTO overtime_requests(member_id,employee_no,member_name,start_date,start_time,end_date,end_time,
        rest_start_time,rest_end_time,overtime_type,reason) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
        (u['id'],employee_no,u['name'],start_date,start_time,end_date,end_time,
         d.get('rest_start_time'),d.get('rest_end_time'),overtime_type,d.get('reason')))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM overtime_requests WHERE id=?",(c.lastrowid,)).fetchone())),201


@overtime_bp.route('/api/overtime/<int:oid>', methods=['PUT'])
@login_required
def upd_overtime(oid):
    u=current_user(); d=request.json or {}; db=get_db()
    row=r2d(db.execute("SELECT * FROM overtime_requests WHERE id=?",(oid,)).fetchone())
    if not row: return '',404
    if not _can_edit_overtime(u,row): return jsonify({'error':'无权限'}),403
    start_date=d.get('start_date') or row['start_date']; start_time=d.get('start_time') or row['start_time']
    end_date=d.get('end_date') or row['end_date']; end_time=d.get('end_time') or row['end_time']
    overtime_type=d.get('overtime_type') or row['overtime_type']
    if overtime_type not in ('转加班费','转调休'):
        return jsonify({'error':'加班类型不合法'}),400
    db.execute("""UPDATE overtime_requests SET start_date=?,start_time=?,end_date=?,end_time=?,
        rest_start_time=?,rest_end_time=?,overtime_type=?,reason=?,updated_at=? WHERE id=?""",
        (start_date,start_time,end_date,end_time,
         d.get('rest_start_time',row.get('rest_start_time')),d.get('rest_end_time',row.get('rest_end_time')),
         overtime_type,d.get('reason',row.get('reason')),now_str(),oid))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM overtime_requests WHERE id=?",(oid,)).fetchone()))


@overtime_bp.route('/api/overtime/<int:oid>', methods=['DELETE'])
@login_required
def del_overtime(oid):
    u=current_user(); db=get_db()
    row=r2d(db.execute("SELECT * FROM overtime_requests WHERE id=?",(oid,)).fetchone())
    if not row: return '',404
    if not _can_edit_overtime(u,row): return jsonify({'error':'无权限'}),403
    db.execute("DELETE FROM overtime_requests WHERE id=?",(oid,)); db.commit()
    return '',204


@overtime_bp.route('/api/overtime/<int:oid>/lock', methods=['POST'])
@admin_required
def lock_overtime(oid):
    d=request.json or {}; db=get_db()
    row=db.execute("SELECT id FROM overtime_requests WHERE id=?",(oid,)).fetchone()
    if not row: return '',404
    db.execute("UPDATE overtime_requests SET locked=?,updated_at=? WHERE id=?",
        (1 if d.get('locked') else 0,now_str(),oid))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM overtime_requests WHERE id=?",(oid,)).fetchone()))


@overtime_bp.route('/api/export/overtime')
@admin_required
def exp_overtime():
    """Export overtime records to Excel. With `month`: single-month, one sheet.
    Without `month` (year only): whole year, one sheet per month (Jan-Dec).
    Group scope: 超级管理员(admin) may pass `groups` (comma-separated group names) to combine
    selected groups into one file, or omit it to export all groups. 普通管理员 are always
    locked to their own group_name, regardless of any `groups` param sent."""
    u=current_user()
    yr=int(request.args.get('year',datetime.date.today().year))
    mo=request.args.get('month')
    is_super=(u.get('username')=='admin')
    if is_super:
        raw_groups=[g.strip() for g in request.args.get('groups','').split(',') if g.strip()]
        group_list=raw_groups  # empty = all groups
        scope_label='全部组' if not group_list else '、'.join(group_list)
    else:
        group_list=[u.get('group_name') or '']
        scope_label=u.get('group_name') or ''
    db=get_db()
    wb=Workbook(); wb.remove(wb.active)
    months=[int(mo)] if mo else list(range(1,13))
    for m in months:
        month_str="{}-{:02d}".format(yr,m)
        sql="""SELECT o.*, m.group_name AS grp FROM overtime_requests o
               JOIN members m ON m.id=o.member_id WHERE o.start_date LIKE ?"""
        params=[month_str+'%']
        if group_list:
            sql+=" AND m.group_name IN ({})".format(','.join('?'*len(group_list)))
            params+=group_list
        sql+=" ORDER BY o.start_date,o.id"
        rows=rs(db.execute(sql,params).fetchall())
        ws=_ot_sheet(wb, "{}年{}月加班".format(str(yr)[-2:],m))
        for i,r in enumerate(rows, start=3):
            _ot_row(ws, i, [
                r.get('employee_no') or '', r.get('member_name') or '',
                (r['start_date'] or '').replace('-','/'), r['start_time'],
                (r['end_date'] or '').replace('-','/'), r['end_time'],
                r.get('rest_start_time') or '', r.get('rest_end_time') or '',
                _OT_TYPE_MAP.get(r.get('overtime_type'), r.get('overtime_type') or ''),
                r.get('reason') or '',
            ])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname="加班记录_{}_{}-{:02d}.xlsx".format(scope_label,yr,int(mo)) if mo else "加班记录_{}_{}.xlsx".format(scope_label,yr)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fname)
