# -*- coding: utf-8 -*-
"""Tasks: CRUD, progress logs (incl. the day/week/month/year progress timeline), Gantt, stats, and Excel export."""
import io, datetime, calendar
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from .db import get_db, r2d, rs
from .utils import today, now_str, auto_risk, STATUS_ZH, TYPE_ZH, current_user, login_required
from .excel import mkhdr, rf, title_cell

tasks_bp = Blueprint('tasks', __name__)


def task_where(u, extra='', extra_params=()):
    if u['is_admin']: w,p="1=1",[]
    else: w,p="group_name=?",[u['group_name']]
    if extra: w+=" AND "+extra; p+=list(extra_params)
    return w,p


def can_edit(u,task):
    if u['is_admin']: return True
    return task.get('assignee_id')==u['id'] or task.get('created_by')==u['id']


@tasks_bp.route('/api/tasks')
@login_required
def all_tasks():
    u=current_user(); db=get_db()
    tt=request.args.get('type')
    w,p=task_where(u,"task_type=?",(tt,)) if tt else task_where(u)
    rows=db.execute("SELECT * FROM tasks WHERE {} ORDER BY updated_at DESC".format(w),p).fetchall()
    return jsonify(rs(rows))


@tasks_bp.route('/api/tasks/mine')
@login_required
def my_tasks():
    u=current_user(); db=get_db()
    tt=request.args.get('type')
    if tt:
        rows=db.execute("SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?) AND task_type=? ORDER BY updated_at DESC",(u['id'],u['id'],tt)).fetchall()
    else:
        rows=db.execute("SELECT * FROM tasks WHERE assignee_id=? OR created_by=? ORDER BY updated_at DESC",(u['id'],u['id'])).fetchall()
    return jsonify(rs(rows))


@tasks_bp.route('/api/tasks/risk')
@login_required
def risk_tasks():
    u=current_user(); db=get_db(); t=today()
    w,p=task_where(u,"status NOT IN ('DELIVERED','CANCELLED','RESOLVED','CLOSED','REJECTED','COMPLETED') AND (plan_end_date<? OR has_risk=1)",(t,))
    return jsonify(rs(db.execute("SELECT * FROM tasks WHERE {} ORDER BY plan_end_date".format(w),p).fetchall()))


@tasks_bp.route('/api/tasks/gantt')
@login_required
def gantt_tasks():
    u=current_user(); db=get_db()
    mid=request.args.get('member_id',u['id'])
    yr=int(request.args.get('year',datetime.date.today().year))
    mo=request.args.get('month')
    if not u['is_admin'] and str(mid)!=str(u['id']): mid=u['id']
    if mo:
        s="{}-{:02d}-01".format(yr,int(mo))
        e="{}-{:02d}-{:02d}".format(yr,int(mo),calendar.monthrange(yr,int(mo))[1])
    else:
        s="{}-01-01".format(yr); e="{}-12-31".format(yr)
    rows=db.execute("""SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?)
        AND plan_start_date IS NOT NULL AND plan_end_date IS NOT NULL
        AND plan_start_date<=? AND plan_end_date>=?
        ORDER BY plan_start_date""",(mid,mid,e,s)).fetchall()
    return jsonify(rs(rows))


@tasks_bp.route('/api/tasks/stats/<gn>')
@login_required
def task_stats(gn):
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    s=request.args.get('startMonth',''); e=request.args.get('endMonth','')
    tt=request.args.get('type','REQUIREMENT')
    rows=rs(get_db().execute("SELECT delivery_month,status,COUNT(*) as cnt FROM tasks WHERE group_name=? AND task_type=? AND delivery_month>=? AND delivery_month<=? GROUP BY delivery_month,status ORDER BY delivery_month",(gn,tt,s,e)).fetchall())
    monthly={}
    for r in rows: monthly.setdefault(r['delivery_month'] or 'unknown',{})[r['status']]=r['cnt']
    return jsonify({'groupName':gn,'startMonth':s,'endMonth':e,'monthlyStats':monthly,'taskType':tt})


@tasks_bp.route('/api/tasks', methods=['POST'])
@login_required
def add_task():
    u=current_user(); d=dict(request.json); db=get_db()
    # Auto-assign to self if not specified
    aid=d.get('assignee_id') or u['id']
    assignee=r2d(db.execute("SELECT id,name,group_name FROM members WHERE id=?",(aid,)).fetchone())
    if assignee and assignee.get('group_name')!=u['group_name'] and not (u['is_admin'] or u.get('can_cross_group')):
        return jsonify({'error':'无权限为其他组成员创建任务'}),403
    d['assignee_id']=aid
    if not d.get('assignee_name'): d['assignee_name']=assignee['name'] if assignee else u['name']
    d['group_name']=(assignee.get('group_name') if assignee else None) or u['group_name']
    d['created_by']=u['id']
    if not d.get('reporter_name'): d['reporter_name']=u['name']
    if not d.get('delivery_month') and d.get('plan_end_date'): d['delivery_month']=d['plan_end_date'][:7]
    d=auto_risk(d)
    c=db.execute("""INSERT INTO tasks(title,description,task_type,status,priority,severity,issue_type,
        assignee_id,assignee_name,reporter_name,group_name,plan_start_date,plan_end_date,actual_end_date,
        delivery_month,progress,has_risk,risk_description,version,module,location,estimated_days,parent_task_id,
        requirement_no,issue_no,created_by)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (d.get('title'),d.get('description'),d.get('task_type','REQUIREMENT'),
         d.get('status','PENDING'),d.get('priority','MEDIUM'),d.get('severity'),d.get('issue_type'),
         d.get('assignee_id'),d.get('assignee_name'),d.get('reporter_name'),d.get('group_name'),
         d.get('plan_start_date'),d.get('plan_end_date'),d.get('actual_end_date'),
         d.get('delivery_month'),d.get('progress',0),d.get('has_risk',0),d.get('risk_description'),
         d.get('version'),d.get('module'),d.get('location'),int(d.get('estimated_days') or 0),d.get('parent_task_id'),
         d.get('requirement_no'),d.get('issue_no'),d.get('created_by')))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM tasks WHERE id=?",(c.lastrowid,)).fetchone())),201


@tasks_bp.route('/api/tasks/<int:tid>', methods=['GET'])
@login_required
def get_task(tid):
    row=get_db().execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone()
    return jsonify(r2d(row)) if row else ('',404)


@tasks_bp.route('/api/tasks/<int:tid>', methods=['PUT'])
@login_required
def upd_task(tid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone())
    if not existing: return ('',404)
    if not can_edit(u,existing): return jsonify({'error':'只能修改自己创建或负责的任务'}),403
    d=dict(request.json)
    # Helper: 如果 d 里没传某字段（key缺失或值为None），fallback 用 existing 的值，避免误清空
    def keep(key, default=None):
        v = d.get(key)
        if v is None or v == '':
            return existing.get(key, default)
        return v
    new_aid=keep('assignee_id')
    assignee=r2d(db.execute("SELECT id,name,group_name FROM members WHERE id=?",(new_aid,)).fetchone()) if new_aid else None
    if assignee and assignee.get('group_name')!=u['group_name'] and not (u['is_admin'] or u.get('can_cross_group')):
        return jsonify({'error':'无权限将任务指派给其他组成员'}),403
    new_group=(assignee.get('group_name') if assignee else None) or existing.get('group_name') or u['group_name']
    if d.get('status') in ('DELIVERED','COMPLETED','RESOLVED','CLOSED') and not d.get('actual_end_date'):
        d['actual_end_date']=today()
        if not d.get('progress'): d['progress']=100
    d=auto_risk(d); db.execute("""UPDATE tasks SET title=?,description=?,task_type=?,status=?,priority=?,
        severity=?,issue_type=?,assignee_id=?,assignee_name=?,reporter_name=?,group_name=?,plan_start_date=?,
        plan_end_date=?,actual_end_date=?,delivery_month=?,progress=?,has_risk=?,risk_description=?,
        version=?,module=?,location=?,estimated_days=?,parent_task_id=?,requirement_no=?,issue_no=?,
        updated_at=? WHERE id=?""",
        (keep('title'),d.get('description'),keep('task_type'),
         keep('status'),keep('priority'),d.get('severity'),d.get('issue_type'),
         new_aid,keep('assignee_name'),d.get('reporter_name'),new_group,
         d.get('plan_start_date'),d.get('plan_end_date'),d.get('actual_end_date'),
         d.get('delivery_month'),d.get('progress',existing.get('progress',0)),
         d.get('has_risk',0),d.get('risk_description'),
         d.get('version'),d.get('module'),d.get('location'),
         int(d.get('estimated_days') or existing.get('estimated_days') or 0),
         keep('parent_task_id'),d.get('requirement_no'),d.get('issue_no'),now_str(),tid))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone()))


@tasks_bp.route('/api/tasks/<int:tid>', methods=['DELETE'])
@login_required
def del_task(tid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone())
    if not existing: return ('',404)
    if not u['is_admin'] and existing.get('created_by')!=u['id']: return jsonify({'error':'只能删除自己创建的任务'}),403
    db.execute("DELETE FROM tasks WHERE id=? OR parent_task_id=?",(tid,tid)); db.commit(); return '',204


# ── Task Logs ─────────────────────────────────────────────────
@tasks_bp.route('/api/tasks/<int:tid>/logs')
@login_required
def get_logs(tid):
    rows=get_db().execute("SELECT * FROM task_logs WHERE task_id=? ORDER BY log_date DESC, id DESC",(tid,)).fetchall()
    return jsonify(rs(rows))


@tasks_bp.route('/api/tasks/<int:tid>/logs', methods=['POST'])
@login_required
def add_log(tid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone())
    if not existing: return ('',404)
    if not can_edit(u,existing): return jsonify({'error':'无权限'}),403
    d=request.json or {}
    content=d.get('content','').strip()
    if not content: return jsonify({'error':'进展内容不能为空'}),400
    log_date=d.get('log_date',today())
    if not u['is_admin']:
        min_date=(datetime.date.today()-datetime.timedelta(days=3)).isoformat()
        if log_date<min_date:
            return jsonify({'error':'进展日期不能早于3天前，如需补录历史进展请联系管理员'}),400
    new_prog=d.get('progress'); new_stat=d.get('status')
    try: hours=max(0.0,float(d.get('hours') or 0))
    except (TypeError,ValueError): hours=0.0
    if new_prog is not None or new_stat is not None:
        upd=[]; prm=[]
        if new_prog is not None: upd.append("progress=?"); prm.append(int(new_prog))
        if new_stat is not None: upd.append("status=?"); prm.append(new_stat)
        upd.append("updated_at=?"); prm.append(now_str()); prm.append(tid)
        db.execute("UPDATE tasks SET {} WHERE id=?".format(','.join(upd)),prm)
    c=db.execute("INSERT INTO task_logs(task_id,member_id,member_name,log_date,content,progress_snapshot,status_snapshot,hours) VALUES(?,?,?,?,?,?,?,?)",
        (tid,u['id'],u['name'],log_date,content,
         new_prog if new_prog is not None else existing.get('progress'),
         new_stat or existing.get('status'),hours))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM task_logs WHERE id=?",(c.lastrowid,)).fetchone())),201


def _can_edit_log(u, log):
    return u['is_admin'] or log['member_id']==u['id']


@tasks_bp.route('/api/tasks/<int:tid>/logs/<int:lid>', methods=['PUT'])
@login_required
def upd_log(tid, lid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM task_logs WHERE id=? AND task_id=?",(lid,tid)).fetchone())
    if not existing: return ('',404)
    if not _can_edit_log(u,existing): return jsonify({'error':'只能修改自己提交的进展记录'}),403
    d=request.json or {}
    content=(d.get('content') if d.get('content') is not None else existing['content']).strip()
    if not content: return jsonify({'error':'进展内容不能为空'}),400
    log_date=d.get('log_date') or existing['log_date']
    if not u['is_admin']:
        min_date=(datetime.date.today()-datetime.timedelta(days=3)).isoformat()
        if log_date<min_date:
            return jsonify({'error':'进展日期不能早于3天前，如需修改历史进展请联系管理员'}),400
    new_prog=d.get('progress',existing.get('progress_snapshot'))
    new_stat=d.get('status',existing.get('status_snapshot'))
    try: hours=max(0.0,float(d.get('hours',existing.get('hours')) or 0))
    except (TypeError,ValueError): hours=existing.get('hours') or 0.0
    db.execute("UPDATE task_logs SET log_date=?,content=?,progress_snapshot=?,status_snapshot=?,hours=? WHERE id=?",
        (log_date,content,new_prog,new_stat,hours,lid))
    db.commit()
    return jsonify(r2d(db.execute("SELECT * FROM task_logs WHERE id=?",(lid,)).fetchone()))


@tasks_bp.route('/api/tasks/<int:tid>/logs/<int:lid>', methods=['DELETE'])
@login_required
def del_log(tid, lid):
    u=current_user(); db=get_db()
    existing=r2d(db.execute("SELECT * FROM task_logs WHERE id=? AND task_id=?",(lid,tid)).fetchone())
    if not existing: return ('',404)
    if not _can_edit_log(u,existing): return jsonify({'error':'只能删除自己提交的进展记录'}),403
    db.execute("DELETE FROM task_logs WHERE id=?",(lid,))
    db.commit()
    return '',204


@tasks_bp.route('/api/tasks/logs/mine')
@login_required
def my_logs():
    u=current_user()
    days=int(request.args.get('days',7))
    since=(datetime.date.today()-datetime.timedelta(days=days)).isoformat()
    rows=get_db().execute("""SELECT l.*,t.title as task_title,t.task_type FROM task_logs l
        JOIN tasks t ON l.task_id=t.id WHERE l.member_id=? AND l.log_date>=?
        ORDER BY l.log_date DESC,l.id DESC LIMIT 50""",(u['id'],since)).fetchall()
    return jsonify(rs(rows))


@tasks_bp.route('/api/tasks/<int:tid>/subtasks')
@login_required
def get_subtasks(tid):
    """Get all subtasks of a parent task."""
    rows=get_db().execute("SELECT * FROM tasks WHERE parent_task_id=? ORDER BY created_at",(tid,)).fetchall()
    return jsonify(rs(rows))


@tasks_bp.route('/api/tasks/by_member')
@login_required
def tasks_by_member():
    """Admin: get tasks grouped by selected members with daily logs."""
    u = current_user()
    if not u['is_admin']: return jsonify({'error':'无权限'}),403
    member_ids = request.args.get('member_ids','')
    date_filter = request.args.get('date')   # optional: YYYY-MM-DD
    gn = request.args.get('group_name', u['group_name'])
    db = get_db()
    mids = [int(x) for x in member_ids.split(',') if x.strip().isdigit()] if member_ids else []
    if not mids:
        rows = rs(db.execute("SELECT id,name FROM members WHERE group_name=? AND is_active=1 ORDER BY id",(gn,)).fetchall())
        mids = [r['id'] for r in rows]
    result = []
    for mid in mids:
        member = r2d(db.execute("SELECT id,name,role,group_name FROM members WHERE id=?",(mid,)).fetchone())
        if not member: continue
        tasks = rs(db.execute("SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?) ORDER BY plan_start_date,created_at",(mid,mid)).fetchall())
        # attach logs
        for t in tasks:
            logs = rs(db.execute("SELECT * FROM task_logs WHERE task_id=? ORDER BY log_date DESC,id DESC",(t['id'],)).fetchall())
            t['logs'] = logs
            # check if today has a log
            t['has_today_log'] = any(l['log_date']==today() for l in logs)
            if date_filter:
                t['has_date_log'] = any(l['log_date']==date_filter for l in logs)
        result.append({'member': member, 'tasks': tasks})
    return jsonify(result)


@tasks_bp.route('/api/tasks/gantt_multi')
@login_required
def gantt_multi():
    """Gantt for multiple members (admin view)."""
    u = current_user()
    if not u['is_admin']: return jsonify({'error':'无权限'}),403
    member_ids = request.args.get('member_ids','')
    yr  = int(request.args.get('year', datetime.date.today().year))
    mo  = request.args.get('month')
    gn  = request.args.get('group_name', u['group_name'])
    db  = get_db()
    mids = [int(x) for x in member_ids.split(',') if x.strip().isdigit()] if member_ids else []
    if not mids:
        rows = rs(db.execute("SELECT id FROM members WHERE group_name=? AND is_active=1",(gn,)).fetchall())
        mids = [r['id'] for r in rows]
    if mo:
        s = "{}-{:02d}-01".format(yr,int(mo))
        e = "{}-{:02d}-{:02d}".format(yr,int(mo),calendar.monthrange(yr,int(mo))[1])
    else:
        s = "{}-01-01".format(yr); e = "{}-12-31".format(yr)
    result = []
    for mid in mids:
        member = r2d(db.execute("SELECT id,name,role,group_name FROM members WHERE id=?",(mid,)).fetchone())
        if not member: continue
        tasks = rs(db.execute("""SELECT * FROM tasks WHERE (assignee_id=? OR created_by=?)
            AND plan_start_date IS NOT NULL AND plan_end_date IS NOT NULL
            AND plan_start_date<=? AND plan_end_date>=?
            ORDER BY plan_start_date""",(mid,mid,e,s)).fetchall())
        for t in tasks:
            logs = rs(db.execute("SELECT log_date,content,progress_snapshot,member_name FROM task_logs WHERE task_id=? ORDER BY log_date",(t['id'],)).fetchall())
            t['logs'] = logs
            t['log_dates'] = [l['log_date'] for l in logs]
        result.append({'member': member, 'tasks': tasks})
    return jsonify({'members': result, 'startDate': s, 'endDate': e})


@tasks_bp.route('/api/tasks/today_todo')
@login_required
def today_todo():
    """Tasks spanning today (active tasks where plan_start<=today<=plan_end)."""
    u = current_user(); db = get_db(); t = today()
    if u['is_admin']:
        where,params = task_where(u,
            "plan_start_date<=? AND plan_end_date>=? AND status NOT IN ('DELIVERED','COMPLETED','RESOLVED','CLOSED','CANCELLED','REJECTED')",
            (t,t))
    else:
        where = "(assignee_id=? OR created_by=?) AND plan_start_date<=? AND plan_end_date>=? AND status NOT IN ('DELIVERED','COMPLETED','RESOLVED','CLOSED','CANCELLED','REJECTED')"
        params = [u['id'],u['id'],t,t]
    rows = rs(db.execute("SELECT * FROM tasks WHERE {} ORDER BY priority DESC,plan_end_date".format(where),params).fetchall())
    for task in rows:
        task['has_today_log'] = db.execute("SELECT 1 FROM task_logs WHERE task_id=? AND log_date=?",(task['id'],t)).fetchone() is not None
    return jsonify(rows)


@tasks_bp.route('/api/tasks/progress')
@login_required
def progress_view():
    """The current user's own task-progress log timeline (day/week/month/year
    browsing is a client-side concern — the caller just picks a [start,end] range).
    Always scoped to self; no cross-member viewing (by design)."""
    u = current_user(); db = get_db()
    start = request.args.get('start') or today()
    end = request.args.get('end') or start
    order = 'ASC' if request.args.get('sort')=='asc' else 'DESC'
    rows = rs(db.execute("""SELECT l.*,t.title as task_title,t.task_type FROM task_logs l
        JOIN tasks t ON l.task_id=t.id
        WHERE l.member_id=? AND l.log_date>=? AND l.log_date<=?
        ORDER BY l.log_date {0},l.id {0}""".format(order),(u['id'],start,end)).fetchall())
    by_type={}; by_date={}; total_hours=0.0
    for r in rows:
        h=r.get('hours') or 0.0
        total_hours+=h
        by_type[r['task_type']]=by_type.get(r['task_type'],0.0)+h
        by_date[r['log_date']]=by_date.get(r['log_date'],0.0)+h
    return jsonify({
        'start':start,'end':end,'logs':rows,
        'totalHours':round(total_hours,2),'totalLogs':len(rows),
        'byType':{k:round(v,2) for k,v in by_type.items()},
        'byDate':{k:round(v,2) for k,v in by_date.items()},
    })


@tasks_bp.route('/api/export/progress')
@login_required
def exp_progress():
    """Export the current user's own progress logs for [start,end] to Excel:
    a raw sorted log sheet + a by-task-type pivot summary sheet."""
    u=current_user(); db=get_db()
    start = request.args.get('start') or today()
    end = request.args.get('end') or start
    order = 'ASC' if request.args.get('sort')=='asc' else 'DESC'
    rows = rs(db.execute("""SELECT l.*,t.title as task_title,t.task_type FROM task_logs l
        JOIN tasks t ON l.task_id=t.id
        WHERE l.member_id=? AND l.log_date>=? AND l.log_date<=?
        ORDER BY l.log_date {0},l.id {0}""".format(order),(u['id'],start,end)).fetchall())
    wb=Workbook()
    ws=wb.active; ws.title="进展明细"
    title_cell(ws,"{} 进展记录（{} ~ {}）".format(u['name'],start,end),8)
    mkhdr(ws,3,['日期','任务标题','类型','进度','状态','工时','内容','记录时间'],[12,32,10,8,10,8,44,18])
    for r in rows:
        ws.append([r['log_date'],r.get('task_title') or '',TYPE_ZH.get(r.get('task_type',''),''),
                   "{}%".format(r['progress_snapshot']) if r.get('progress_snapshot') is not None else '',
                   STATUS_ZH.get(r.get('status_snapshot',''),r.get('status_snapshot') or ''),
                   r.get('hours') or 0,r.get('content') or '',r.get('created_at') or ''])
    by_type={}
    for r in rows: by_type[r['task_type']]=by_type.get(r['task_type'],0.0)+(r.get('hours') or 0.0)
    ws2=wb.create_sheet(title="分类汇总")
    title_cell(ws2,"按事务类型汇总工时",2)
    mkhdr(ws2,3,['事务类型','工时（小时）'],[16,16])
    for k,v in by_type.items():
        ws2.append([TYPE_ZH.get(k,k),round(v,2)])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name="{}_进展记录_{}_{}.xlsx".format(u['name'],start,end))


@tasks_bp.route('/api/export/tasks/<gn>')
@login_required
def exp_tasks(gn):
    u=current_user()
    if not u['is_admin'] and u['group_name']!=gn: return jsonify({'error':'无权限'}),403
    tt=request.args.get('type','REQUIREMENT')
    s=request.args.get('startMonth',''); e=request.args.get('endMonth','')
    rows=rs(get_db().execute("SELECT * FROM tasks WHERE group_name=? AND task_type=? AND delivery_month>=? AND delivery_month<=? ORDER BY delivery_month",(gn,tt,s,e)).fetchall())
    tname=TYPE_ZH.get(tt,tt); wb=Workbook(); ws=wb.active; ws.title=tname+"报表"
    title_cell(ws,"{} {} ({} ~ {})".format(gn,tname,s,e),12)
    mkhdr(ws,3,['标题','类型','模块','版本','负责人','优先级','状态','进度','计划开始','计划结束','实际完成','风险'],
          [32,10,12,10,10,8,10,8,12,12,12,22])
    for r in rows:
        risk=('⚠ '+(r.get('risk_description') or '有风险')) if r.get('has_risk') else '正常'
        ws.append([r['title'],TYPE_ZH.get(r.get('task_type',''),''),r.get('module',''),r.get('version',''),
                   r.get('assignee_name',''),r.get('priority',''),STATUS_ZH.get(r.get('status',''),r.get('status','')),
                   "{}%".format(r.get('progress',0)),r.get('plan_start_date',''),r.get('plan_end_date',''),r.get('actual_end_date',''),risk])
        if r.get('has_risk'):
            for col in range(1,13): ws.cell(row=ws.max_row,column=col).fill=rf()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name="{}_{}_{}_{}.xlsx".format(gn,tname,s,e))
