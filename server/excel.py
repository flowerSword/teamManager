# -*- coding: utf-8 -*-
"""Shared openpyxl formatting helpers used by every /api/export/* route."""
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def mkhdr(ws,row,hdrs,wds):
    f=PatternFill("solid",fgColor="1F3864"); ft=Font(bold=True,color="FFFFFF",size=11); al=Alignment(horizontal="center",vertical="center")
    for i,h in enumerate(hdrs,1):
        c=ws.cell(row=row,column=i,value=h); c.fill=f; c.font=ft; c.alignment=al
        ws.column_dimensions[get_column_letter(i)].width=wds[i-1]


def rf(): return PatternFill("solid",fgColor="FFE0B2")


def title_cell(ws,text,cols):
    ws.merge_cells('A1:{}1'.format(get_column_letter(cols)))
    ws['A1'].value=text; ws['A1'].font=Font(bold=True,size=14); ws['A1'].alignment=Alignment(horizontal='center')
    ws.row_dimensions[1].height=28
